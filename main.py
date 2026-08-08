import requests
import copy
import time
import openpyxl
import os
import sys
import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog, scrolledtext
import threading
import queue
import json
from datetime import datetime
import urllib3
import uuid  # 新增：用于生成唯一设备ID

# 禁用 SSL 警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ==================== 日志控制台类 ====================

class LogConsole:
    """日志控制台窗口，支持 info 和 error 分级输出"""

    def __init__(self, title="执行日志控制台"):
        self.root = tk.Tk()
        self.root.title(title)
        self.root.geometry("900x600")
        self.root.configure(bg="#1e1e1e")
        self.root.minsize(600, 400)

        # 顶部状态栏
        self.status_frame = tk.Frame(self.root, bg="#2d2d2d", height=35)
        self.status_frame.pack(fill=tk.X, side=tk.TOP)
        self.status_frame.pack_propagate(False)

        self.status_label = tk.Label(
            self.status_frame,
            text="⏳ 等待开始...",
            bg="#2d2d2d",
            fg="#cccccc",
            font=("Microsoft YaHei", 10),
            anchor=tk.W,
            padx=10
        )
        self.status_label.pack(fill=tk.BOTH, expand=True)

        # 统计信息栏
        self.stats_frame = tk.Frame(self.root, bg="#252526", height=30)
        self.stats_frame.pack(fill=tk.X, side=tk.TOP)
        self.stats_frame.pack_propagate(False)

        self.info_count = 0
        self.error_count = 0
        self.success_count = 0

        self.stats_label = tk.Label(
            self.stats_frame,
            text="📊 Info: 0 | ❌ Error: 0 | ✅ Success: 0",
            bg="#252526",
            fg="#aaaaaa",
            font=("Consolas", 9),
            anchor=tk.W,
            padx=10
        )
        self.stats_label.pack(fill=tk.BOTH, expand=True)

        # 日志文本区域
        self.text_area = scrolledtext.ScrolledText(
            self.root,
            wrap=tk.WORD,
            bg="#1e1e1e",
            fg="#d4d4d4",
            font=("Consolas", 10),
            padx=10,
            pady=10,
            borderwidth=0,
            highlightthickness=0
        )
        self.text_area.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        # 配置标签颜色
        self.text_area.tag_configure("timestamp", foreground="#858585")
        self.text_area.tag_configure("info", foreground="#d4d4d4")
        self.text_area.tag_configure("info_highlight", foreground="#4ec9b0")
        self.text_area.tag_configure("error", foreground="#f44747", background="#3a1d1d")
        self.text_area.tag_configure("error_highlight", foreground="#ff6b6b")
        self.text_area.tag_configure("success", foreground="#4ec9b0")
        self.text_area.tag_configure("warning", foreground="#dcdcaa")
        self.text_area.tag_configure("separator", foreground="#3c3c3c")

        # 底部按钮栏
        self.btn_frame = tk.Frame(self.root, bg="#2d2d2d", height=40)
        self.btn_frame.pack(fill=tk.X, side=tk.BOTTOM)
        self.btn_frame.pack_propagate(False)

        self.clear_btn = tk.Button(
            self.btn_frame,
            text="🗑️ 清空日志",
            command=self.clear,
            bg="#3c3c3c",
            fg="#cccccc",
            activebackground="#505050",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            font=("Microsoft YaHei", 9),
            cursor="hand2"
        )
        self.clear_btn.pack(side=tk.LEFT, padx=10, pady=5)

        self.save_btn = tk.Button(
            self.btn_frame,
            text="💾 保存日志",
            command=self.save_log,
            bg="#3c3c3c",
            fg="#cccccc",
            activebackground="#505050",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            font=("Microsoft YaHei", 9),
            cursor="hand2"
        )
        self.save_btn.pack(side=tk.LEFT, padx=5, pady=5)

        self.top_btn = tk.Button(
            self.btn_frame,
            text="⬆️ 返回顶部",
            command=self.scroll_to_top,
            bg="#3c3c3c",
            fg="#cccccc",
            activebackground="#505050",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            font=("Microsoft YaHei", 9),
            cursor="hand2"
        )
        self.top_btn.pack(side=tk.LEFT, padx=5, pady=5)

        self.bottom_btn = tk.Button(
            self.btn_frame,
            text="⬇️ 滚动到底部",
            command=self.scroll_to_bottom,
            bg="#3c3c3c",
            fg="#cccccc",
            activebackground="#505050",
            activeforeground="#ffffff",
            relief=tk.FLAT,
            font=("Microsoft YaHei", 9),
            cursor="hand2"
        )
        self.bottom_btn.pack(side=tk.LEFT, padx=5, pady=5)

        # 消息队列，用于线程间通信
        self.msg_queue = queue.Queue()
        self._check_queue()

        # 绑定关闭事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.is_closing = False
        self.worker_thread = None

        # 初始日志
        self._do_log_info("日志控制台已启动", "system")
        self._do_log_info("等待用户输入参数...", "system")

    def _get_timestamp(self):
        return datetime.now().strftime("%H:%M:%S.%f")[:-3]

    def _update_stats(self):
        self.stats_label.config(
            text=f"📊 Info: {self.info_count} | ❌ Error: {self.error_count} | ✅ Success: {self.success_count}"
        )

    def _check_queue(self):
        """定时检查消息队列，从工作线程接收日志"""
        try:
            while True:
                msg = self.msg_queue.get_nowait()
                msg_type = msg.get('type', 'info')
                if msg_type == 'info':
                    self._do_log_info(msg['message'], msg.get('category', 'general'))
                elif msg_type == 'error':
                    self._do_log_error(msg['message'], msg.get('exception'))
                elif msg_type == 'success':
                    self._do_log_success(msg['message'])
                elif msg_type == 'warning':
                    self._do_log_warning(msg['message'])
                elif msg_type == 'separator':
                    self._do_log_separator(msg.get('char', '='), msg.get('count', 50))
                elif msg_type == 'status':
                    self._do_set_status(msg['text'], msg.get('status_type', 'normal'))
                elif msg_type == 'messagebox':
                    self._do_show_messagebox(msg['title'], msg['message'])
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self._check_queue)

    # ========== 线程安全的日志方法 ==========

    def log_info(self, message, category="general"):
        """线程安全的 info 日志"""
        self.msg_queue.put({'type': 'info', 'message': message, 'category': category})

    def log_error(self, message, exception=None):
        """线程安全的 error 日志"""
        exc_str = str(exception) if exception else None
        self.msg_queue.put({
            'type': 'error',
            'message': message,
            'exception': exc_str
        })

    def log_success(self, message):
        """线程安全的 success 日志"""
        self.msg_queue.put({'type': 'success', 'message': message})

    def log_warning(self, message):
        """线程安全的 warning 日志"""
        self.msg_queue.put({'type': 'warning', 'message': message})

    def log_separator(self, char="=", count=50):
        """线程安全的 separator"""
        self.msg_queue.put({'type': 'separator', 'char': char, 'count': count})

    def set_status(self, text, status_type="normal"):
        """线程安全的设置状态"""
        self.msg_queue.put({
            'type': 'status',
            'text': text,
            'status_type': status_type
        })

    def show_messagebox(self, title, message):
        """线程安全的弹窗"""
        self.msg_queue.put({
            'type': 'messagebox',
            'title': title,
            'message': message
        })

    # ========== 实际的UI操作方法（只能在主线程调用） ==========

    def _do_log_info(self, message, category="general"):
        """实际的 info 日志输出（必须在主线程）"""
        if self.is_closing:
            return
        self.info_count += 1
        self._update_stats()
        timestamp = self._get_timestamp()

        self.text_area.configure(state=tk.NORMAL)
        self.text_area.insert(tk.END, f"[{timestamp}] ", "timestamp")

        category_colors = {
            "system": "info_highlight",
            "http": "info_highlight",
            "excel": "info_highlight",
            "general": "info"
        }
        tag = category_colors.get(category, "info")
        self.text_area.insert(tk.END, f"[{category.upper()}] ", tag)
        self.text_area.insert(tk.END, f"{message}\n", "info")
        self.text_area.see(tk.END)
        self.text_area.configure(state=tk.DISABLED)

    def _do_log_error(self, message, exception=None):
        """实际的 error 日志输出（必须在主线程）"""
        if self.is_closing:
            return
        self.error_count += 1
        self._update_stats()
        timestamp = self._get_timestamp()

        self.text_area.configure(state=tk.NORMAL)
        self.text_area.insert(tk.END, "─" * 80 + "\n", "separator")
        self.text_area.insert(tk.END, f"[{timestamp}] ", "timestamp")
        self.text_area.insert(tk.END, "[ERROR] ", "error_highlight")
        self.text_area.insert(tk.END, f"{message}\n", "error")
        if exception:
            self.text_area.insert(tk.END, f"    ↳ 异常: {exception}\n", "error")
        self.text_area.insert(tk.END, "─" * 80 + "\n", "separator")
        self.text_area.see(tk.END)
        self.text_area.configure(state=tk.DISABLED)
        self.status_label.config(text=f"❌ 发生错误: {message[:50]}...", fg="#f44747")

    def _do_log_success(self, message):
        self.success_count += 1
        self._update_stats()
        timestamp = self._get_timestamp()
        self.text_area.configure(state=tk.NORMAL)
        self.text_area.insert(tk.END, f"[{timestamp}] ", "timestamp")
        self.text_area.insert(tk.END, "[SUCCESS] ", "success")
        self.text_area.insert(tk.END, f"{message}\n", "success")
        self.text_area.see(tk.END)
        self.text_area.configure(state=tk.DISABLED)

    def _do_log_warning(self, message):
        timestamp = self._get_timestamp()
        self.text_area.configure(state=tk.NORMAL)
        self.text_area.insert(tk.END, f"[{timestamp}] ", "timestamp")
        self.text_area.insert(tk.END, "[WARN] ", "warning")
        self.text_area.insert(tk.END, f"{message}\n", "warning")
        self.text_area.see(tk.END)
        self.text_area.configure(state=tk.DISABLED)

    def _do_log_separator(self, char="=", count=50):
        self.text_area.configure(state=tk.NORMAL)
        self.text_area.insert(tk.END, f"\n{char * count}\n\n", "separator")
        self.text_area.see(tk.END)
        self.text_area.configure(state=tk.DISABLED)

    def _do_set_status(self, text, status_type="normal"):
        colors = {
            "normal": "#cccccc",
            "success": "#4ec9b0",
            "error": "#f44747",
            "warning": "#dcdcaa",
            "running": "#569cd6"
        }
        self.status_label.config(text=text, fg=colors.get(status_type, "#cccccc"))

    def _do_show_messagebox(self, title, message):
        """在主线程显示弹窗"""
        messagebox.showinfo(title, message)

    def start_worker(self, target, args=()):
        """在后台线程启动工作函数"""
        self.worker_thread = threading.Thread(target=target, args=args, daemon=True)
        self.worker_thread.start()

    def clear(self):
        """清空日志"""
        self.text_area.configure(state=tk.NORMAL)
        self.text_area.delete(1.0, tk.END)
        self.text_area.configure(state=tk.DISABLED)
        self.info_count = 0
        self.error_count = 0
        self.success_count = 0
        self._update_stats()

    def save_log(self):
        """保存日志到文件"""
        try:
            filepath = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
                initialfile=f"log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            if filepath:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(self.text_area.get(1.0, tk.END))
                self.log_success(f"日志已保存到: {filepath}")
        except Exception as e:
            self.log_error("保存日志失败", e)

    def scroll_to_top(self):
        self.text_area.see("1.0")

    def scroll_to_bottom(self):
        self.text_area.see(tk.END)

    def on_closing(self):
        self.is_closing = True
        if messagebox.askokcancel("确认退出", "确定要关闭日志控制台吗？\n关闭后将终止程序执行。"):
            self.root.destroy()
            sys.exit(0)

    def run(self):
        """启动控制台主循环"""
        self.root.mainloop()


# ==================== 全局日志实例 ====================

console = None


def get_base_dir():
    """获取程序运行时的基础目录（支持打包后）"""
    if getattr(sys, 'frozen', False):
        if hasattr(sys, '_MEIPASS'):
            return os.path.dirname(sys.executable)
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def send_sms(login_phone):
    """发送验证码"""
    send_code_url = "https://jt.jinmaodigital.com/api/identity/appuser/SendCode"
    send_code_headers = copy.deepcopy(headers)
    # 修复1：使用动态生成的唯一设备ID
    send_code_headers["deviceid"] = DEVICE_ID
    send_code_dict = {"phoneNumber": login_phone}

    try:
        # 修复2：使用 session 发送请求，自动管理 Cookie
        response = session.post(
            send_code_url,
            json=send_code_dict,
            headers=send_code_headers,
            timeout=10
        )
        return True
    except Exception as e:
        return False


def get_user_input():
    """获取用户输入参数（在主线程运行）"""
    root = tk.Tk()
    root.withdraw()

    choice = messagebox.askyesno("选择方式", "是否通过文件对话框选择Excel？\n选'否'则手动输入路径")

    if choice:
        excel_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        print(f"通过文件对话框选择: {excel_path}")
    else:
        excel_path = simpledialog.askstring(
            "输入Excel路径",
            "请输入Excel文件路径:\n（支持相对路径，相对于exe所在目录）",
            initialvalue="和达HC).xlsx"
        )
        print(f"手动输入路径: {excel_path}")

    if not excel_path:
        messagebox.showerror("错误", "必须输入Excel路径")
        sys.exit(1)

    excel_path = excel_path.strip().strip('"').strip("'")

    if not os.path.isabs(excel_path):
        base_dir = get_base_dir()
        excel_path = os.path.join(base_dir, excel_path)
        print(f"[调试] 相对路径转换为绝对路径: {excel_path}")

    source_level1 = simpledialog.askstring(
        "输入一级菜单",
        "请输入客户来源的一级菜单名称:",
        initialvalue="自渠-基础渠道"
    )
    if not source_level1:
        messagebox.showerror("错误", "必须输入一级菜单")
        sys.exit(1)

    source_level2 = simpledialog.askstring(
        "输入二级菜单",
        "请输入客户来源的二级菜单名称:\n（没有则留空或填'无'）",
        initialvalue="电call"
    )

    # 新增：登录手机号输入
    login_phone = simpledialog.askstring(
        "登录手机号",
        "请输入接收验证码的登录手机号：",
        initialvalue=""
    )
    if not login_phone:
        messagebox.showerror("错误", "登录手机号不能为空")
        sys.exit(1)
    login_phone = login_phone.strip()

    # 新增：客户备注输入
    customer_remark = simpledialog.askstring(
        "客户备注",
        "请输入报备统一客户备注（可为空）：",
        initialvalue=""
    )
    customer_remark = customer_remark.strip()

    send_success = send_sms(login_phone)
    if not send_success:
        messagebox.showerror("错误", "验证码发送失败，请检查网络")
        sys.exit(1)

    auth_code = simpledialog.askstring(
        "输入验证码",
        "请输入手机验证码:",
        initialvalue=""
    )

    confirm = messagebox.askyesno(
        "确认信息",
        f"Excel: {excel_path}\n一级菜单: {source_level1}\n二级菜单: {source_level2 or '无'}\n登录手机号:{login_phone}\n客户备注:{customer_remark}\n验证码: {auth_code}\n确认开始？"
    )
    if not confirm:
        print("用户取消操作")
        sys.exit(0)

    root.destroy()
    return excel_path, source_level1.strip(), source_level2.strip() if source_level2 else "", auth_code, login_phone, customer_remark


def process_with_retry(user_name, user_phone, access_token, source_level1, source_level2, remark, user_info):
    """发送报备请求"""
    report_url = "https://jt.jinmaodigital.com/api/customer/NewAddClue"
    token = f"Bearer {access_token}"
    report_headers = copy.deepcopy(headers)
    report_headers["Authorization"] = token
    # 修复1：使用动态生成的唯一设备ID，替代原来的硬编码
    report_headers["deviceid"] = DEVICE_ID
    raw_projects = user_info.get("appProjects", [])
    current_projects = [{"projectId": p["projectId"], "projectName": p["projectName"], "sort": p.get("sort", 0)} for p
                        in raw_projects]

    report_dict = {
        "name": user_name,
        "phoneNumber": user_phone,
        "internationalCode": "",
        "isHiddenPhoneNumber": False,
        "gender": 0,
        "teamId": user_info["teams"][0]["id"],
        "projects": current_projects,
        "commissionerId": user_info.get("id", ''),
        "propertyConsultant": "",
        "intentionCode": 20,
        "intentionCodeStr": "中",
        "remark": remark,
        "customType": 2,
        "ytkMemberId": user_info.get("ytkMemberId", ''),
        "cognitionWayFirst": "400e5c27-3d03-48b3-aa41-3742c1814e12",
        "cognitionWayFirstStr": "自渠团队-自获客",
        "cognitionWaySecond": "08d5edb4-840a-46c3-b4d4-1fab7e7dfcda",
        "cognitionWaySecondStr": source_level1,
        "cognitionWayThird": "e3c3acf9-f2c7-4b12-a74f-9bbed48cf069",
        "cognitionWayThirdStr": source_level2
    }

    try:
        console.log_info(f"发送报备请求: {user_name} | {user_phone}", "http")
        # 修复2：使用 session 发送请求，自动管理 Cookie
        report_res = session.post(
            report_url,
            json=report_dict,
            headers=report_headers,
            timeout=15
        )
        console.log_info(f"HTTP状态码: {report_res.status_code}", "http")
        console.log_info(f"响应内容: {report_res.text[:300]}", "http")
        report_data = report_res.json().get("data", "")
        return report_data
    except Exception as e:
        console.log_error(f"报备请求失败: {user_name} | {user_phone}", e)
        raise


def main_worker():
    """后台工作线程的主函数"""
    global EXCEL_PATH, SOURCE_LEVEL1, SOURCE_LEVEL2, AUTH_CODE, LOGIN_PHONE, CUSTOMER_REMARK

    console.set_status("🔐 正在登录认证...", "running")

    auth_url = "https://jt.jinmaodigital.com/api/identity/appuserdetail/AppLogin"
    auth_headers = copy.deepcopy(headers)
    auth_dict = {
        "phoneNumber": LOGIN_PHONE,
        "code": AUTH_CODE,
        # 修复1：使用动态生成的唯一设备ID，替代原来的硬编码
        "registrationID": DEVICE_ID,
        "platform": "android"
    }

    try:
        # 修复2：使用 session 发送请求，自动管理 Cookie
        auth_res = session.post(
            auth_url,
            json=auth_dict,
            headers=auth_headers,
            timeout=10
        )
        console.log_info(f"认证响应: {auth_res.text[:300]}", "http")

        auth_json = auth_res.json()
        console.log_info(f"完整用户信息: {json.dumps(auth_json, ensure_ascii=False)}")
        access_token = auth_json["access_token"]
        user_info = auth_json["user"]

        console.log_success("登录认证成功")
        console.log_info(f"用户信息: {user_info.get('name', 'Unknown')}", "system")
    except Exception as e:
        console.log_error("登录认证失败", e)
        return

    console.set_status("📖 正在读取Excel...", "running")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        total_rows = ws.max_row
        console.log_success(f"Excel加载成功: {EXCEL_PATH}")
        console.log_info(f"总数据行数: {total_rows}", "excel")
    except Exception as e:
        console.log_error("Excel加载失败", e)
        return

    console.set_status("🚀 开始处理数据...", "running")
    console.log_separator("=", 60)
    console.log_info("开始批量报备处理", "system")
    console.log_separator("=", 60)

    for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
        phone_cell = '未知手机号'
        try:
            name_cell = row[0]
            phone_cell = row[1]

            name = str(name_cell.value).strip() if name_cell.value else ""
            phone = str(phone_cell.value).strip() if phone_cell.value else ""

            if not name or not phone:
                console.log_warning(f"第 {idx} 行数据不完整，跳过: name={name}, phone={phone}")
                continue

            console.log_separator("─", 50)
            console.log_info(f"📋 开始处理第 {name_cell.row} 行: {name} | {phone}", "excel")
            console.set_status(f"🔄 处理中: {name} ({idx}/{total_rows})", "running")

            report_res_str = process_with_retry(
                user_name=name,
                user_phone=phone,
                access_token=access_token,
                source_level1=SOURCE_LEVEL1,
                source_level2=SOURCE_LEVEL2,
                remark=CUSTOMER_REMARK,
                user_info=user_info
            )

            ws.cell(row=name_cell.row, column=3, value=report_res_str)
            wb.save(EXCEL_PATH)

            console.log_success(f"✅ 第 {name_cell.row} 行处理成功")

        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            console.log_error(f"❌ 第 {idx} 行报备失败: {phone_cell}", e)
            console.log_info(f"详细错误:\n{error_trace}", "error")
        finally:
            console.log_info(f"⏱️ 等待10秒后处理下一行...", "system")
            time.sleep(10)

    console.log_separator("=", 60)
    console.log_success("🎉 全部处理完成！")
    console.set_status("✅ 全部处理完成", "success")

    # 使用队列在主线程显示完成弹窗
    console.show_messagebox("完成", "全部处理完成！")
    wb.close()


# ==================== 程序入口 ====================

if __name__ == '__main__':
    # ==================== 修复核心开始 ====================
    # 1. 每个进程启动时生成唯一的设备ID（20位十六进制字符串）
    #    A窗口和B窗口会生成不同的ID，服务器会识别为两台不同设备
    DEVICE_ID = uuid.uuid4().hex[:20]

    # 2. 创建全局 Session 对象，自动管理 Cookie
    #    彻底移除硬编码的 acw_tc，让服务器通过响应头 Set-Cookie 自动设置
    session = requests.Session()
    session.verify = False  # 全局禁用 SSL 验证（替代原来每个请求里的 verify=False）

    # 可选：先访问一次网站首页，让 WAF 下发初始 Cookie（如 acw_tc）
    # 如果发送验证码时遇到 WAF 拦截，取消下面这行的注释
    # try:
    #     session.get("https://jt.jinmaodigital.com", timeout=5)
    # except Exception:
    #     pass
    # ==================== 修复核心结束 ====================

    # 全局headers（已移除硬编码的 cookie）
    headers = {
        "accept": "application/json",
        "content-type": "application/json",
        "accept-encoding": "gzip",
        "user-agent": "okhttp/4.12.0"
        # 修复：移除硬编码 "cookie": "acw_tc=..."，由 session 自动管理
    }

    # 第1步：弹出输入对话框（主线程）
    EXCEL_PATH, SOURCE_LEVEL1, SOURCE_LEVEL2, AUTH_CODE, LOGIN_PHONE, CUSTOMER_REMARK = get_user_input()

    print(f"Excel: {EXCEL_PATH}")
    print(f"一级菜单: {SOURCE_LEVEL1}")
    print(f"二级菜单: {SOURCE_LEVEL2 or '无'}")
    print(f"登录手机号: {LOGIN_PHONE}")
    print(f"客户备注: {CUSTOMER_REMARK}")
    print(f"验证码: {AUTH_CODE}")
    print(f"本进程设备ID: {DEVICE_ID}")  # 调试用，可看到每个窗口ID不同

    # 第2步：初始化日志控制台
    console = LogConsole("报备系统 - 执行日志控制台")

    console.log_info(f"最终配置:", "system")
    console.log_info(f"  Excel: {EXCEL_PATH}", "system")
    console.log_info(f"  一级菜单: {SOURCE_LEVEL1}", "system")
    console.log_info(f"  二级菜单: {SOURCE_LEVEL2 or '无'}", "system")
    console.log_info(f"  登录手机号: {LOGIN_PHONE}", "system")
    console.log_info(f"  客户备注: {CUSTOMER_REMARK}", "system")
    console.log_info(f"  验证码: {AUTH_CODE}", "system")
    console.log_info(f"  设备ID: {DEVICE_ID}", "system")  # 日志中显示本窗口设备ID

    # 第3步：在后台线程启动业务逻辑
    console.start_worker(main_worker)

    # 第4步：主线程运行 Tkinter 事件循环（保持界面响应）
    console.run()
